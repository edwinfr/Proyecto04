import streamlit as st
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

st.title("Ventas")

candidatos = [Path("c.xlsx"), Path("datos_ejemplo.xlsx")]
archivo = next((p for p in candidatos if p.exists()), None)

if archivo is None:
    st.error("No se encontró ningún archivo Excel en la carpeta del proyecto.")
else:
    try:
        libro = load_workbook(archivo, data_only=True)
        hoja = libro["Facturas por venta"]

        rows = list(hoja.iter_rows(values_only=True))
        data_rows = [row for row in rows if any(cell is not None for cell in row)]

        if len(data_rows) < 2:
            st.info("No hay datos suficientes para mostrar.")
        else:
            headers = ["Fecha", "Factura", "Monto"]
            valores = []
            for row in data_rows:
                if row[2] == "Fecha" and row[3] == "Factura" and row[4] == "Monto":
                    continue
                if row[2] is not None and row[3] is not None and row[4] is not None:
                    valores.append([row[2], row[3], row[4]])

            datos = pd.DataFrame(valores, columns=headers)
            datos["Monto"] = pd.to_numeric(datos["Monto"], errors="coerce")
            total_montos = float(datos["Monto"].sum())

            st.subheader(f"Datos de {archivo.name} - Facturas por venta")
            st.dataframe(datos)
            st.success(f"Total de montos: {total_montos:,.2f}")
    except Exception as e:
        st.error(f"No se pudo leer el archivo Excel: {e}")